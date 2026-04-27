from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse


def build_forum_dashboard_excel(data):
    wb = Workbook()

    # --------------------------------------------------
    # Sheet 1: KPIs
    # --------------------------------------------------
    ws = wb.active
    ws.title = "KPIs"

    ws.append([
        "Key",
        "Label",
        "Value",
        "Change Percent",
        "Positive Trend"
    ])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in data["kpis"]:
        ws.append([
            row["key"],
            row["label"],
            row["value"],
            row["change_percent"],
            row["isPositive"],
        ])

    # --------------------------------------------------
    # Sheet 2: Forum Activity Trend
    # --------------------------------------------------
    ws2 = wb.create_sheet("Activity Trend")

    ws2.append([
        "Date",
        "Posts",
        "Comments"
    ])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for row in data["forum_activity_trend"]:
        ws2.append([
            row["date"],
            row["posts"],
            row["comments"],
        ])

    # --------------------------------------------------
    # Sheet 3: Top Categories
    # --------------------------------------------------
    ws3 = wb.create_sheet("Top Categories")

    ws3.append([
        "Category",
        "Count"
    ])

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for row in data["top_discussion_categories"]:
        ws3.append([
            row["name"],
            row["count"],
        ])

    # --------------------------------------------------
    # Auto Column Width
    # --------------------------------------------------
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            max_length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = max_length + 3

    return wb


def workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response