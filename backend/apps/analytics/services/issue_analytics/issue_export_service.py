from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse

def build_issue_dashboard_excel(data):
    wb = Workbook()

    # --------------------------------------------------
    # Sheet 1: Summary
    # --------------------------------------------------
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Metric", "Value"])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    stats = data["stats"]

    ws.append(["Total Issues", stats["total_issues"]])
    ws.append(["Resolved Issues", stats["resolved_issues"]])
    ws.append(["Pending Issues", stats["pending_issues"]])
    ws.append(["Rejected Issues", stats["rejected_issues"]])

    # --------------------------------------------------
    # Sheet 2: Trend
    # --------------------------------------------------
    ws2 = wb.create_sheet("Trend")
    ws2.append(["Label", "Reported", "Resolved"])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for row in data["trend_chart"]:
        ws2.append([
            row["label"],
            row["reported"],
            row["resolved"]
        ])

    # --------------------------------------------------
    # Sheet 3: Funnel
    # --------------------------------------------------
    ws3 = wb.create_sheet("Funnel")
    ws3.append(["Stage", "Count"])

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for row in data["funnel_chart"]:
        ws3.append([
            row["stage"],
            row["count"]
        ])

    # --------------------------------------------------
    # Sheet 4: Zones
    # --------------------------------------------------
    ws4 = wb.create_sheet("Zones")
    ws4.append(["Zone", "Issues"])

    for cell in ws4[1]:
        cell.font = Font(bold=True)

    for row in data["zone_chart"]:
        ws4.append([
            row["zone"],
            row["issues"]
        ])

    # --------------------------------------------------
    # Sheet 5: Categories
    # --------------------------------------------------
    ws5 = wb.create_sheet("Categories")
    ws5.append(["Category", "Issues"])

    for cell in ws5[1]:
        cell.font = Font(bold=True)

    for row in data["category_chart"]:
        ws5.append([
            row["name"],
            row["value"]
        ])

    # --------------------------------------------------
    # Auto width
    # --------------------------------------------------
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = length + 3

    return wb




def workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response