from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse


def build_poll_dashboard_excel(data):
    wb = Workbook()

    # --------------------------------------------------
    # Sheet 1: Summary KPI
    # --------------------------------------------------
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Metric", "Value", "Change %"])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    kpis = data["kpis"]

    ws.append([
        "Total Polls",
        kpis["total_polls"]["value"],
        kpis["total_polls"]["change_percent"]
    ])

    ws.append([
        "Total Votes",
        kpis["total_votes"]["value"],
        kpis["total_votes"]["change_percent"]
    ])

    ws.append([
        "Participation Rate",
        kpis["participation_rate"]["value"],
        kpis["participation_rate"]["change_percent"]
    ])

    ws.append([
        "Avg Votes Per Poll",
        kpis["avg_votes_per_poll"]["value"],
        kpis["avg_votes_per_poll"]["change_percent"]
    ])

    # --------------------------------------------------
    # Sheet 2: Participation Trend
    # --------------------------------------------------
    ws2 = wb.create_sheet("Trend")

    ws2.append([
        "Date",
        "Votes",
        "Unique Voters",
        "Active Polls"
    ])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for row in data["participation_trend"]:
        ws2.append([
            row["date"],
            row["votes"],
            row["voters"],
            row["active_polls"]
        ])

    # --------------------------------------------------
    # Sheet 3: Top Polls
    # --------------------------------------------------
    ws3 = wb.create_sheet("Top Polls")

    ws3.append([
        "Poll ID",
        "Title",
        "Votes",
        "Participants",
        "Status"
    ])

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for row in data["top_polls"]:
        ws3.append([
            row["id"],
            row["title"],
            row["votes"],
            row["unique_participants"],
            row["status"]
        ])

    # --------------------------------------------------
    # Auto Column Width
    # --------------------------------------------------
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            max_length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = max_length + 3

    return wb


def workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response