from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse


def build_user_dashboard_excel(data):
    wb = Workbook()

    # --------------------------------------------------
    # Sheet 1: Summary
    # --------------------------------------------------
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Metric", "Value"])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    stats = data["stats"]

    ws.append(["Total Users", stats["total_users"]])
    ws.append(["New Users This Month", stats["new_users_this_month"]])
    ws.append(["Active Users", stats["active_users"]])
    ws.append(["Citizens", stats["citizens"]])
    ws.append(["Solvers", stats["solvers"]])
    ws.append(["Admins", stats["admins"]])

    # --------------------------------------------------
    # Sheet 2: Distribution
    # --------------------------------------------------
    ws2 = wb.create_sheet("Distribution")

    ws2.append(["Role", "Count"])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for row in data["distribution"]:
        ws2.append([
            row["name"],
            row["value"],
        ])

    # --------------------------------------------------
    # Sheet 3: Growth
    # --------------------------------------------------
    ws3 = wb.create_sheet("Growth")

    ws3.append(["Date", "New Users"])

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for row in data["growth"]:
        ws3.append([
            row["date"],
            row["users"],
        ])

    # --------------------------------------------------
    # Sheet 4: Zone Solvers
    # --------------------------------------------------
    ws4 = wb.create_sheet("Zone Solvers")

    ws4.append([
        "Zone",
        "Total Solvers",
        "Active Solvers"
    ])

    for cell in ws4[1]:
        cell.font = Font(bold=True)

    for row in data["zone_solver_chart"]:
        ws4.append([
            row["zone"],
            row["solvers"],
            row["active_solvers"],
        ])

    # --------------------------------------------------
    # Auto Width
    # --------------------------------------------------
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )

            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = length + 3

    return wb


def workbook_response(workbook, filename):
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response