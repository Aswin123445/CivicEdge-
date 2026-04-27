from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def build_volunteer_dashboard_excel(data):
    wb = Workbook()

    # --------------------------------------------------
    # Sheet 1: KPIs
    # --------------------------------------------------
    ws = wb.active
    ws.title = "KPIs"

    ws.append(["Metric", "Value", "Change (%)"])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for key, value in data["kpis"].items():
        ws.append([
            key.replace("_", " ").title(),
            value["value"],
            value["change_percent"]
        ])

    # --------------------------------------------------
    # Sheet 2: Growth
    # --------------------------------------------------
    ws2 = wb.create_sheet("Growth")
    ws2.append(["Month", "Joined"])

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for row in data["growth"]:
        ws2.append([
            row["month"],
            row["joined"]
        ])

    # --------------------------------------------------
    # Sheet 3: Group Access Distribution
    # --------------------------------------------------
    ws3 = wb.create_sheet("Access Distribution")
    ws3.append(["Type", "Count"])

    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for row in data["group_access_distribution"]:
        ws3.append([
            row["name"],
            row["value"]
        ])

    # --------------------------------------------------
    # Sheet 4: Top Participation Groups
    # --------------------------------------------------
    ws4 = wb.create_sheet("Top Groups")
    ws4.append(["Group Name", "Participants"])

    for cell in ws4[1]:
        cell.font = Font(bold=True)

    for row in data["top_participation_groups"]:
        ws4.append([
            row["group"],
            row["participants"]
        ])

    # --------------------------------------------------
    # Sheet 5: Conversion Funnel
    # --------------------------------------------------
    ws5 = wb.create_sheet("Funnel")
    ws5.append(["Stage", "Count"])

    for cell in ws5[1]:
        cell.font = Font(bold=True)

    for row in data["conversion_funnel"]:
        ws5.append([
            row["stage"],
            row["count"]
        ])

    # --------------------------------------------------
    # Auto column width
    # --------------------------------------------------
    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = length + 3

    return wb